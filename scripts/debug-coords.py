#!/usr/bin/env python3
"""Debug: qué filas fallan en parseo de coords y por qué."""
import re
from pathlib import Path

import openpyxl

EXCEL = Path(__file__).resolve().parent / "Listado_RAF_Repetidoras.xlsx"


def dms_to_decimal(s, is_south_or_west=False):
    if s is None or (isinstance(s, (int, float)) and not isinstance(s, bool)):
        return float(s) if s is not None else None
    s = str(s).strip().replace(",", ".")
    m = re.search(r"(-?\d+)\s*[^\d\w]+\s*(\d+)\s*[^\d\w]*\s*([\d.]+)", s)
    if not m:
        return None
    d, mi = int(m.group(1)), int(m.group(2))
    sec = float(m.group(3))
    if mi > 59:
        mi = mi % 60
    if sec >= 60:
        sec = sec % 60
    dec = d + mi / 60 + sec / 3600
    if is_south_or_west and dec > 0:
        dec = -dec
    return round(dec, 10)


def main():
    wb = openpyxl.load_workbook(EXCEL, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    i_sig = 3  # Signal before Tx
    i_lat = headers.index("Latitud") if "Latitud" in headers else 13
    i_lon = headers.index("Longitud") if "Longitud" in headers else 14

    failed = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        sig = str(row[i_sig] or "").strip()
        lat_raw = row[i_lat] if i_lat < len(row) else None
        lon_raw = row[i_lon] if i_lon < len(row) else None
        lat = dms_to_decimal(lat_raw, True)
        lon = dms_to_decimal(lon_raw, True) if lon_raw else None
        if lat is None or lon is None:
            failed.append({
                "row": row_idx,
                "signal": sig,
                "lat_raw": repr(lat_raw),
                "lon_raw": repr(lon_raw),
                "lat_parsed": lat,
                "lon_parsed": lon,
            })

    wb.close()
    print(f"Filas que fallan: {len(failed)}")
    for f in failed[:30]:
        print(f"  Row {f['row']:4} {f['signal']:25} | lat={f['lat_raw']:30} lon={f['lon_raw']}")


if __name__ == "__main__":
    main()
